import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import signal
import os
from openpyxl import Workbook

# 创建一个新的Excel工作簿
workbook = Workbook()

# 读取数据
dir_name = 'E:/psdforturb'
file_list = os.listdir(dir_name)

for file_name in file_list:
    if file_name.endswith('.csv'):
        total_name = os.path.join(dir_name, file_name)
        data_temp = pd.read_csv(total_name, skiprows=119, usecols=[1])
        data = data_temp.values.squeeze()

        # 参数设置
        fs = 20000  # 采样频率
        fmin = 0.01    # 最小通过频率
        fmax = 999 # 最大通过频率
        filterorder = 2  # 滤波器阶数

        # 滤波和功率谱密度估计
        n_blocks = 4  # 数据分块数量
        n_samples = len(data)
        pxx_list = []
        freq_list = []

        for i in range(n_blocks):
            start = i * n_samples // n_blocks
            end = (i + 1) * n_samples // n_blocks
            xx = data[start:end]

            # 滤波
            nyquist = fs / 2
            low = fmin / nyquist
            high = fmax / nyquist
            b, a = signal.butter(filterorder, [low, high], btype='band', analog=False)
            sig = signal.filtfilt(b, a, xx)

            # 功率谱密度估计
        nfft = int(np.ceil(fs / fmin))
        for i in range(n_blocks):
            start = i * n_samples // n_blocks
            end = (i + 1) * n_samples // n_blocks
            xx = data[start:end]

            # 滤波
            nyquist = fs / 2
            low = fmin / nyquist
            high = fmax / nyquist
            b, a = signal.butter(filterorder, [low, high], btype='band', analog=False)
            sig = signal.filtfilt(b, a, xx)

            # 修正 nfft，确保不会超过信号长度
            nfft = min(int(np.ceil(fs / fmin)), len(sig))
            window = signal.get_window('hann', nfft)

            # 功率谱密度估计
            f, pxx = signal.welch(sig, fs, window=window, nperseg=nfft)
            pxx_list.append(pxx)
            freq_list.append(f)

        # 将功率谱密度估计结果存储到Excel工作表中
        worksheet = workbook.create_sheet(title=file_name.split('.')[0])
        worksheet['A1'] = 'Frequency'
        worksheet['B1'] = 'PSD_Block1'
        worksheet['C1'] = 'PSD_Block2'
        worksheet['D1'] = 'PSD_Block3'
        worksheet['E1'] = 'PSD_Block4'

        for i in range(len(freq_list[0])):
            worksheet.cell(row=i+2, column=1, value=freq_list[0][i])
            for j in range(n_blocks):
                worksheet.cell(row=i+2, column=j+2, value=pxx_list[j][i])

        # # 绘图
        # pxx = np.array(pxx_list).T
        # freq = freq_list[0]
        # plt.figure()
        # plt.loglog(freq, pxx)
        # plt.xlabel('频率/Hz')
        # plt.ylabel('PSD/(vertically shifted)')
        # plt.title(file_name)
        # plt.show()

# 保存Excel工作簿
workbook.save('psd_results2.xlsx')
